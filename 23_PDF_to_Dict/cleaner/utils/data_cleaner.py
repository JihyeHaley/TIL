from nltk.tokenize import sent_tokenize
import stanza
import pandas as pd
import logging
import re
import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.utils.dataframe import dataframe_to_rows


global url_col
global keys

url_col = False
keys = []

# Downloads language models
# stanza.download('en')
# stanza.download('ko')
# stanza.download('ja')
# stanza.download('zh')

# STANZA Support Languages
SUPPORT_LANGUAGE = ['zh', 'ja', 'ko', 'en']


def pick_cols(df, keylist):
    global url_col
    global keys

    keys = keylist
    if df.keys()[-1] == 'url':
        url_col = True
    if url_col:
        picks = keylist + ['url']
    else:
        picks = keylist
    return df[picks].drop_duplicates(subset=keylist), url_col, keys


def separate_loner(df, lang, drop=True):    # 다른 컬럼에 아무런 정보 없이 언어 하나에만 데이터가 있는 경우, 경로가 잘못되어 짝을 잃어버린 데이터라고 가정
    global keys
    loner = df.copy()
    for key in keys:
        if key == lang: continue
        loner = loner[(loner[key].isna()) | (loner[key] == '[]')]
    if drop:
        df = df.drop(loner.index)
    return df, loner


def flatten(df, str_type=True):    # <input type : pd.DataFrame>
    global url_col
    global keys
    if url_col:
        flattened = {i: [] for i in (keys+['url'])}
    else:
        flattened = {i: [] for i in keys}

    if type(df.iloc[0,0]) == list:
        str_type = False

    if str_type:
        df = df.fillna('[]')
        for col in keys:
            df[col] = df[col].map(lambda x: eval(x))

    df['size'] = df.iloc[:, 0].map(lambda x: len(x))
    for idx in range(len(df)):
        line = df.iloc[idx]
        if url_col:
            flattened['url'].extend([line['url']]*line['size'])

        for lang in keys:
            standard = line['size']
            diff = standard - len(line[lang])  # 각 셀 안의 리스트 안에 있는 원소의 개수
            if diff > 0:  # 한국어의 데이터 개수와 맞지 않을 경우 'NoneFlag'를 동일한 개수로 넣어줌
                if type(line[lang]) == list:
                    line[lang].extend([None for i in range(diff)])
                else:
                    line[lang] = ([None for i in range(standard)])
            flattened[lang].extend(line[lang][:standard])  # 만들어놓은 flatten dictionary에 데이터 추가
    return pd.DataFrame(flattened).drop_duplicates().reset_index(drop=True)


def rmv_ugly(df, lang, ugly_pattern):  # lang : col_name
    ugly_data = []
    for idx, text in enumerate(df[lang]):
        if not text: continue
        find = re.search(ugly_pattern, text)
        if find and find.group():
            ugly_data.append(idx)
    return df.drop(df.index[ugly_data])


def rmv_single_words(df, ko):  # ko : korean col_name
    ugly_data = []
    for idx, text in enumerate(df[ko]):
        if not text: continue
        find = re.search(r'\s.*\s', text)    # 한국어로 공백이 최소 2회 있어야 문장으로 올바른 문장으로 취급
        if find and find.group():
            pass
        else:
            ugly_data.append(idx)
    return df.drop(df.index[ugly_data])


def unify_quotes(df):
    global keys
    for key in keys:
        df[key] = df[key].map(lambda x: re.sub(r'[“”„‟❝❞⹂〝〞〟]', '"', x))
        df[key] = df[key].map(lambda x: re.sub(r"[‘’‛❛❜]", "'", x))
    return df


def capitalize(text):
    for idx in range(len(text)):
        if text[idx].isalpha():
            return text[:idx] + text[idx].upper() + text[idx+1:]
        else:
            return text


def tokenize_sentence(text, lang_code):
    if lang_code not in SUPPORT_LANGUAGE:
        return sent_tokenize(text)
    tokenizer = stanza.Pipeline(lang=lang_code, processors='tokenize')
    doc = tokenizer(text)
    return [sentence.text for sentence in doc.sentences]


def highlight_none(data): # <input type : pd.DataFrame>
    global url_col
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row in dataframe_to_rows(data, index=False, header=True):
        try: worksheet.append(row)
        except: logging.debug('{}'.format(row))
            # worksheet.append([re.sub(escape_codes, '', i) for i in row]) # 에러가 나는 데이터(2~3줄의 정보)를 살려보려고 했지만... 코드가 듣지를 않

    fill = PatternFill(patternType='solid', fgColor=Color('00FFFF00')) # 조건에 맞는 셀 노란색으로 하이라이트
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    if not url_col:
        max_row += 1
        max_col += 1

    for row in range(1, max_row):
        for column in range(1, max_col):
            cell = worksheet.cell(row=row, column=column)
            if cell.value == 'NoneFlag':    # 해당 시트 내의 데이터 완전 탐색하며 'NoneFlag'인 데이 노란색으로 하이라이트
                cell.fill = fill
                cell.value = None      # 해당 셀을 비우고 싶으면 활성화시킬 것
    return workbook
